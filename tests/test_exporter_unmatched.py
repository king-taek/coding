"""Excel exporter — 매칭 + 미매칭 reference 통합 출력 검증."""

import os
import tempfile
from pathlib import Path

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PyQt6.QtWidgets")
pytest.importorskip("openpyxl")
pytest.importorskip("PIL.Image")

from PIL import Image                                          # noqa: E402
from PyQt6.QtWidgets import QApplication                       # noqa: E402

from aoi_verification.app.models.result import (               # noqa: E402
    FinalResult, MatchResult, MissEntry,
)
from aoi_verification.app.workers.exporter import ExcelExporter  # noqa: E402


@pytest.fixture(scope="module")
def qapp():
    return QApplication.instance() or QApplication([])


def _make_image(folder: Path, name: str) -> Path:
    p = folder / name
    Image.new("RGB", (400, 300), color=(80, 120, 200)).save(str(p), "JPEG")
    return p


def test_export_writes_matches_and_unmatched(qapp, isolated_cache, tmp_path):
    """매칭 + 미매칭이 같이 들어가면 결과 워크북에 두 종류 모두 나타난다."""
    src = tmp_path / "src"
    src.mkdir()
    a_ref = _make_image(src, "a_ref.jpeg")
    a_val = _make_image(src, "a_val.jpeg")
    b_ref = _make_image(src, "b_ref.jpeg")  # 미매칭

    result = FinalResult(
        mode="single",
        ref_machine="1호기",
        val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=a_ref, val_path=a_val,
                              score=0.9, direction="A→B")],
        unmatched_refs=[MissEntry(slot="S1", side="ref", path=b_ref,
                                   note="미매칭")],
    )
    dst = tmp_path / "out.xlsx"
    # 양식 자동 감지를 피하려고 일부러 존재하지 않는 템플릿 경로를 명시 지정.
    no_tpl = tmp_path / "no_template.xlsx"
    exp = ExcelExporter(result, dst_path=dst, template_path=no_tpl)
    exp.run()  # QThread.run() 을 동기 실행
    assert dst.exists()

    from openpyxl import load_workbook
    wb = load_workbook(str(dst))
    ws = wb.active
    # 헤더는 row 1~2 (양식.xlsx 형식), 데이터는 row 3 부터 시작.
    # 정렬: a_ref.jpeg, b_ref.jpeg → 매칭이 먼저, 미매칭이 그 다음 행.
    # row 3: a → 매칭 (D 셀 비어있고 이미지 임베드만)
    # row 4: b → 미매칭 (D 셀에 파일명 텍스트)
    d3 = ws["D3"].value
    d4 = ws["D4"].value
    assert d3 is None, f"매칭 행의 D 셀은 비어있어야 (이미지만 임베드): {d3}"
    assert d4 == "b_ref.jpeg", f"미매칭 행의 D 셀에 파일명: {d4}"

    # 빨간 폰트 확인
    font = ws["D4"].font
    assert font.color is not None
    # openpyxl 의 color 는 ARGB hex 또는 indexed.
    color_str = str(font.color.rgb or font.color.value or "")
    assert "FF2D55" in color_str.upper()

    # 코멘트 ‘미매칭’ 확인
    assert ws["D4"].comment is not None
    assert "미매칭" in str(ws["D4"].comment.text)

    # AOI-N 헤더가 row 2 에 들어갔는지 (#3).
    assert ws["C2"].value == "AOI-1"
    assert ws["D2"].value == "AOI-2"


def test_export_no_unmatched_unaffected(qapp, isolated_cache, tmp_path):
    """unmatched_refs 가 비어있으면 기존 동작과 동일하게 매칭만 출력."""
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "ref.jpeg")
    val = _make_image(src, "val.jpeg")

    result = FinalResult(
        mode="single",
        ref_machine="1호기",
        val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val,
                              score=0.9, direction="A→B")],
    )
    dst = tmp_path / "out2.xlsx"
    no_tpl = tmp_path / "no_template.xlsx"
    ExcelExporter(result, dst_path=dst, template_path=no_tpl).run()
    assert dst.exists()

    from openpyxl import load_workbook
    ws = load_workbook(str(dst)).active
    # 헤더 2 줄 + 데이터 row 3 부터. 행 3 = 매칭, 행 4 는 데이터 없음.
    assert ws["B3"].value == "S1"
    assert ws["D3"].value is None  # 이미지 임베드만, 텍스트 없음
    assert ws["B4"].value is None
