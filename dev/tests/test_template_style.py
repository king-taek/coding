"""엑셀 양식 서식 — 열 폭·배경이 출력에서 살아남는가.

배경(실측 사고 두 가지):

1. **폭이 매번 되돌려졌다.**  ``_do_export`` 가 C~H 를 전부 같은 폭으로 통일하고
   A·B 에 하한(6·14)을 걸어서, 양식에 아무리 좁게 잡아 둬도 출력에서 원래대로
   돌아왔다.  '양식을 고쳤는데 결과가 안 바뀐다' 는 형태라 원인을 찾기 어렵다.

2. **21행부터 배경이 끊겼다.**  배경을 템플릿의 미리 칠한 20행에만 의존하면,
   결함이 20건을 넘는 순간(실제로는 100건 이상) 그 아래는 흰 바탕이 된다.
   ``_style_data_row`` 가 행마다 다시 칠하는 이유다.

여기서는 **실제 `_do_export` 를 돌린 결과 파일**을 열어서 확인한다 — 상수만
비교하면 위 두 사고를 못 잡는다(상수는 맞는데 출력이 다른 게 문제였다).
"""

import os
from pathlib import Path

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PyQt6.QtWidgets")
pytest.importorskip("openpyxl")
pytest.importorskip("PIL.Image")

from PIL import Image                                            # noqa: E402
from PyQt6.QtWidgets import QApplication                         # noqa: E402

from aoi_verification.app.models.result import (                 # noqa: E402
    FinalResult, MatchResult,
)
from aoi_verification.app.utils import paths                     # noqa: E402
from aoi_verification.app.workers import exporter as ex          # noqa: E402

SHEET = ex.SHEET_FULL_NAME
N_ROWS = 25          # 템플릿이 미리 칠해 둔 20행보다 **많게**


@pytest.fixture(scope="module")
def qapp():
    return QApplication.instance() or QApplication([])


@pytest.fixture(scope="module")
def exported(qapp, tmp_path_factory):
    """실제 exporter 를 한 번 돌려 만든 결과 파일(모듈 내 공유)."""
    tmp = tmp_path_factory.mktemp("tplstyle")
    src = tmp / "src"
    src.mkdir()
    shots = []
    for i in range(N_ROWS * 2):
        p = src / f"D{i:03d}.jpg"
        Image.new("RGB", (120, 120), (90, 90, 90)).save(str(p), "JPEG")
        shots.append(p)

    matches = [
        MatchResult(slot=f"slot {i // 9 + 1:02d}",
                    ref_path=shots[i * 2], val_path=shots[i * 2 + 1],
                    score=0.9)
        for i in range(N_ROWS)
    ]
    result = FinalResult(mode="cross", ref_machine="17", val_machine="23",
                         matches=matches)
    dst = tmp / "out.xlsx"
    ExcelExporter = ex.ExcelExporter
    ExcelExporter(result=result, dst_path=dst,
                  template_path=paths.template_path(),
                  include_full_template=True)._do_export()

    import openpyxl
    return openpyxl.load_workbook(dst)


def test_column_widths_survive_the_export(exported):
    """★ 양식이 정한 폭이 그대로 나온다 — 사고 1."""
    ws = exported[SHEET]
    for col, want in ex.COL_WIDTHS.items():
        got = ws.column_dimensions[col].width
        assert got == pytest.approx(want, abs=0.6), (
            f"{col}열 폭이 {got} — 양식이 정한 {want} 로 나와야 한다")


def test_photo_columns_stay_equal(exported):
    """C·D 는 사진이 나란히 들어가므로 서로 같아야 한다."""
    ws = exported[SHEET]
    assert (ws.column_dimensions["C"].width
            == pytest.approx(ws.column_dimensions["D"].width, abs=0.01))


def test_manual_columns_are_narrower_than_photo_columns(exported):
    """수기 열이 사진 열만큼 넓어지면 가로가 다시 길어진다."""
    ws = exported[SHEET]
    photo = ws.column_dimensions["C"].width
    for col in "EFGH":
        assert ws.column_dimensions[col].width < photo


def test_body_style_continues_past_the_template_rows(exported):
    """★ 21행 이후에도 배경·격자가 이어진다 — 사고 2."""
    ws = exported[SHEET]
    last = ex.DATA_START_ROW + N_ROWS - 1
    assert last > 22, "템플릿 미리칠한 범위를 넘겨야 의미 있는 검사다"
    for row in (ex.DATA_START_ROW, 23, last):
        for col in "ABCDEFGH":
            cell = ws[f"{col}{row}"]
            assert cell.fill.fgColor.rgb in ex.BODY_FILLS[col], (
                f"{col}{row} 배경이 {cell.fill.fgColor.rgb} — 그룹 배경이 아니다")
            assert cell.border.left.style == "thin", f"{col}{row} 격자 없음"


def test_banding_alternates_by_data_order(exported):
    """줄무늬는 데이터 순번을 따른다 — 홀짝이 번갈아야 한다."""
    ws = exported[SHEET]
    first, second = ex.DATA_START_ROW, ex.DATA_START_ROW + 1
    assert (ws[f"A{first}"].fill.fgColor.rgb
            != ws[f"A{second}"].fill.fgColor.rgb)


def test_group_colors_differ_between_camtek_and_kla(exported):
    """Camtek(E·F)과 KLA(G·H)가 같은 색이면 색 구분의 의미가 없다."""
    ws = exported[SHEET]
    row = ex.DATA_START_ROW
    assert ws[f"E{row}"].fill.fgColor.rgb != ws[f"G{row}"].fill.fgColor.rgb
    assert ws["E2"].fill.fgColor.rgb != ws["G2"].fill.fgColor.rgb


def test_committed_template_matches_the_exporter_constants():
    """``dev/양식.xlsx`` 는 exporter 상수로 구운 산출물이다.

    손으로 고쳐 두 벌이 갈라지면 '빈 양식' 과 '실제 출력' 의 생김새가 달라진다.
    어긋나면 ``python scripts/internal/make_template.py`` 를 다시 돌려야 한다.
    """
    import openpyxl

    tpl = paths.template_path()
    assert tpl and Path(tpl).is_file(), "양식.xlsx 를 찾지 못했다"
    ws = openpyxl.load_workbook(tpl).active
    for col, want in ex.COL_WIDTHS.items():
        assert ws.column_dimensions[col].width == pytest.approx(want, abs=0.01), (
            f"{col}열 폭이 상수와 다르다 — make_template.py 를 다시 돌려라")
    assert ws["A1"].fill.fgColor.rgb == ex.HEADER_NAVY
    for col, want in ex.GROUP_FILLS.items():
        assert ws[f"{col}2"].fill.fgColor.rgb == want
