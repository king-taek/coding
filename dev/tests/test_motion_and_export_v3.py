"""구조 개편 21~31 — 모션 6건 · 엑셀/결과 4건의 회귀 가드.

※ 27안(시트 발원점 스케일)은 사용자 결정으로 **롤백**했다 — 시트 등장은 예전대로
   화면 하단에서 18px 상승 + 페이드다(`sheet_host._enter`).

시안 머리말이 못박은 예산: **상시 애니메이션 0 · 200~300ms · 기존 motion 헬퍼 원칙
준수.**  그래서 여기서 검사하는 것은 '움직인다' 가 아니라 셋이다:

1. **한 번만 움직인다** — 등장/피드백은 1회성이고 루프가 없다.  상시 애니메이션은
   이 저장소가 '로딩이 버벅인다' 로 이미 한 번 겪은 실수다(회전 링 제거).
2. **레이아웃을 건드리지 않는다** — 대상이 전부 레이아웃이 자리를 정하는 위젯이라,
   `move()`/마진으로 밀면 다음 패스가 되돌리거나 이웃이 함께 출렁인다.  그래서
   이동은 `QGraphicsEffect`(그리기 단계)에서만 한다.
3. **끝나면 흔적을 남기지 않는다** — 이펙트가 붙어 있으면 Qt 가 그 위젯을 계속
   오프스크린으로 다시 그린다.  수백 행에 남기면 스크롤이 무거워진다.

엑셀·결과 4건은 '출력물과 화면이 스스로 설명하는가' 를 본다 — 미매칭이 눈으로
세어지는가, 스크롤·인쇄에서 머리 행이 살아 있는가, 저장이 무엇을 하는 중인지
말하는가, 누르기 전에 목적지를 아는가.
"""

from __future__ import annotations

import os

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import inspect
import tempfile
from pathlib import Path

import pytest

from aoi_verification.app import i18n

pytest.importorskip("PyQt6.QtWidgets")

from PyQt6.QtWidgets import QLabel, QWidget                    # noqa: E402

from aoi_verification.app.ui import motion                     # noqa: E402


@pytest.fixture
def motion_on(monkeypatch):
    """헤드리스에서도 실제 애니메이션 경로를 태운다.

    `motion.enabled()` 는 offscreen 에서 False 라, 켜지 않으면 이 파일의 검사가
    전부 '아무것도 안 함' 경로만 확인하게 된다."""
    monkeypatch.setattr(motion, "enabled", lambda: True)


# ═══ 공통 — 시안이 밀리초로 정한 값들 ════════════════════════════════════════
def test_designed_durations_are_not_scaled():
    """시안이 눈으로 정한 지속시간은 `motion_scale` 을 타지 않는다.

    ★ `dur()` 는 0.8 을 곱한다.  220ms 지정이 176ms 로 나가면 '지정한 값' 과
    '실제 값' 이 갈라진다 — DUR_SHEET/LOADING/RECOLOR 가 이미 그 이유로 예외인
    자리이고, 21·23·25·26안의 값도 같은 성격이다."""
    assert (motion.DUR_RAIL_LEAD, motion.DUR_FINISH_TICK) == (140, 200)
    assert (motion.DUR_RISE_IN, motion.STAGGER_RISE_MS) == (220, 60)
    assert (motion.DUR_SWIPE_OUT, motion.DUR_SWIPE_IN) == (180, 120)
    assert motion.DUR_KNOB == 180
    for fn in (motion._run_offset_fade,):
        # 주석·독스트링은 빼고 **코드 줄**만 본다(설명에 dur() 가 등장한다).
        code = [ln for ln in inspect.getsource(fn).splitlines()
                if not ln.strip().startswith(("#", "★", '"'))]
        assert any("setDuration(int(duration))" in ln for ln in code), (
            f"{fn.__name__} 이 지정값을 그대로 쓰지 않는다")
        assert not any("motion.dur(" in ln or "= dur(" in ln for ln in code), (
            f"{fn.__name__} 이 지정값에 스케일을 곱한다")


def test_offset_fade_moves_without_touching_the_layout(styled_qapp, motion_on):
    """등장/퇴장 이동은 **그리기 단계**에서만 일어난다.

    ★ 검토 행은 QVBoxLayout, 선별 사진은 QScrollArea 안이다.  `move()` 로 밀면
    다음 레이아웃 패스가 즉시 되돌리고, 마진으로 밀면 sizeHint 가 바뀌어 이웃
    행들이 함께 출렁인다 — 26안이 '레이아웃 불변 · 리플로 0' 으로 못박은 성질이
    이것이다."""
    src = inspect.getsource(motion._run_offset_fade)
    assert "setGraphicsEffect" in src
    assert ".move(" not in src, "이동을 위젯 위치로 하고 있다 — 레이아웃과 싸운다"
    assert "setContentsMargins" not in src, "마진으로 밀면 이웃이 함께 밀린다"

    host = QWidget()
    lab = QLabel("x", host)
    host.resize(200, 60)
    lab.setGeometry(10, 10, 100, 20)
    before = lab.geometry()
    motion.rise_in(lab)
    host.grab()                                   # draw() 를 실제로 태운다
    assert lab.geometry() == before, "위젯의 실제 기하가 움직였다"


def test_effect_is_detached_when_the_animation_ends(styled_qapp, motion_on):
    """재생이 끝나면 이펙트를 뗀다 — 붙여 두면 그 위젯이 계속 오프스크린 렌더다."""
    host = QWidget()
    lab = QLabel("x", host)
    anim = motion.rise_in(lab)
    assert lab.graphicsEffect() is not None, "재생 중에는 붙어 있어야 한다"
    anim.setCurrentTime(anim.totalDuration())      # 끝까지 감는다
    assert lab.graphicsEffect() is None, "끝났는데 이펙트가 남아 있다"


def test_a_finished_animation_never_clears_a_newer_effect(styled_qapp,
                                                          motion_on):
    """연타 가드 — 옛 애니메이션이 끝나며 **새로 건 이펙트**를 지우면 안 된다.

    선별 화면의 →/← 연타가 정확히 이 경로다: 사진이 그대로 멈춰 보인다."""
    host = QWidget()
    lab = QLabel("x", host)
    first = motion.rise_in(lab)
    motion.fade_in(lab)                            # 새 등장이 덮어쓴다
    newest = lab.graphicsEffect()
    first.setCurrentTime(first.totalDuration())    # 옛 애니가 뒤늦게 끝난다
    assert lab.graphicsEffect() is newest, "새 이펙트가 지워졌다"


def test_headless_still_calls_back(styled_qapp):
    """모션이 꺼져 있어도 `on_done` 은 온다 — 안 오면 고스트가 영원히 남는다."""
    seen = []
    host = QWidget()
    motion.swipe_out(QLabel("x", host), on_done=lambda: seen.append(1))
    assert seen == [1]


# ═══ 21안-A — 레일 선행 릴레이 ═══════════════════════════════════════════════
def test_rail_fills_first_then_the_page_slides(styled_qapp, motion_on):
    """레일이 **먼저** 채워지고(140ms), 화면이 뒤이어 슬라이드-인(240ms).

    ★ 겹쳐 재생하면 같은 정보('어디로 가는가')를 두 모션이 동시에 말해 서로를
    흐린다 — 순서가 곧 인과(레일 → 화면)를 만든다.  순차라 동시 애니는 늘 1개,
    총 380ms 로 예산 안이다."""
    from aoi_verification.app.ui.widgets.journey_rail import JourneyRail

    rail = JourneyRail()
    rail.set_current(1, animate=True)
    anim = rail._fill_anim
    assert anim is not None, "레일 채우기가 시작되지 않았다"
    assert anim.duration() == motion.DUR_RAIL_LEAD
    # ★ 채움은 **등속**이다(시안 `dsFill .14s linear`).  진행을 나타내는 채움에
    #   감속을 씌우면 '거의 다 찼는데 안 끝나는' 것처럼 읽힌다 — 이 저장소가
    #   결정형 진행바에 등속을 쓰는 이유와 같다.  감속은 본문 슬라이드의 몫이다.
    from PyQt6.QtCore import QEasingCurve
    assert anim.easingCurve().type() == QEasingCurve.Type.Linear
    # 화면 전환은 그 시간만큼 **미뤄진다**(창이 릴레이의 두 번째 박자를 쥔다).
    from aoi_verification.app.ui import main_window as mw
    src = inspect.getsource(mw.MainWindow._show_page)
    assert "DUR_RAIL_LEAD" in src and "_run_pending_page_transition" in src
    rail.deleteLater()


# ═══ 23안-B — 완료의 마침 신호 ═══════════════════════════════════════════════
def test_reaching_100_percent_turns_the_panel_pass_coloured(styled_qapp):
    """마침 신호는 **색**이다(모션 0) — 100% 에 닿으면 statusPass 로 전환된다."""
    from aoi_verification.app.ui.widgets.loading_overlay import LoadingOverlay

    host = QWidget()
    ov = LoadingOverlay(host)
    ov.set_progress(5, 10, "계산 중")
    assert not ov._pct_label.property("state")
    ov.set_progress(10, 10, "계산 중")
    # ★ 색이 바뀌는 것은 **문구 하나**다.  B 목업을 실측하면 pass 는 "유사도 계산
    #   완료" 스팬에만 걸려 있고 눈금은 accent, "100 %" 는 기본 잉크다 — 셋을 다
    #   칠하면 '한 화면에 강조 하나' 가 무너져 완료가 경고처럼 커진다.
    assert ov._label.property("state") == "done", "완료인데 문구 색이 그대로다"
    assert not ov._progress.property("state"), "눈금까지 칠했다"
    assert not ov._pct_label.property("state"), "퍼센트까지 칠했다"
    # 다시 진행 중이 되면 완료 표시는 사라진다(다음 작업이 초록으로 시작하면 안 된다).
    ov.set_progress(0, 0, "다음 단계")
    assert not ov._label.property("state")
    host.deleteLater()


def test_only_the_finish_tick_colours_the_rule(styled_qapp):
    """눈금이 완료색이 되는 것은 **200ms 틱 동안뿐**이다(A안의 그 틱).

    ★ 평상시 100% 는 문구만 초록이다(B).  '수 분 작업의 끝' 한 지점에서만
    눈금이 한 번 빛나고, 오버레이가 걷히면 되돌아간다 — 다음 작업이 완료색으로
    시작하면 그 색이 신호이기를 그만둔다."""
    from aoi_verification.app.ui.widgets.loading_overlay import LoadingOverlay

    host = QWidget()
    ov = LoadingOverlay(host)
    ov.set_progress(10, 10, "계산 중")
    assert not ov._progress.property("state")
    ov.finish_tick()
    assert ov._progress.property("state") == "done", "틱인데 눈금이 그대로다"
    ov._finish_hide()
    assert not ov._progress.property("state"), "틱이 끝났는데 눈금이 초록이다"
    host.deleteLater()


# ═══ 25안 — 검토 리스트 진입 스태거 ══════════════════════════════════════════
def test_entrance_stagger_matches_the_designed_budget():
    """가시 8행 · 60ms 간격 · 220ms 라이즈 — 총 640ms.

    ★ 600행 전체에 걸면 화면 밖 행까지 비용을 내고 마지막 행이 수십 초 뒤에
    나타난다.  스태거가 말하려는 것은 '목록이 방금 만들어졌다 · 여기부터 봐라'
    이지 '행이 600개다' 가 아니다."""
    from aoi_verification.app.ui.pages import match_review_page as mrp

    assert mrp.MatchReviewPage._ENTRANCE_ROWS == 8
    assert mrp.MatchReviewPage._ENTRANCE_STEP_MS == motion.STAGGER_RISE_MS
    total = (mrp.MatchReviewPage._ENTRANCE_ROWS - 1) * motion.STAGGER_RISE_MS \
        + motion.DUR_RISE_IN
    assert total <= 700, f"스태거 전체가 {total}ms — 등장이 아니라 대기가 된다"


def test_entrance_waits_for_the_overlay_and_for_the_page_to_be_shown():
    """스태거는 **사용자가 실제로 목록을 보는 순간** 재생된다.

    ★ 두 번 어긋났던 자리다.  (a) `hide_overlay()` 는 동기 종료가 아니다 —
    최소표시 래치와 페이드아웃이 남아 있으면 타이머만 걸고 돌아오므로, 바로 다음
    줄에서 재생하면 아직 덮개 아래다.  (b) 행 생성은 `load_state` 안에서 끝나는데
    창은 그 **뒤에** 페이지를 스택에 올린다 — 그 시점엔 화면 밖이라 아무도 못 본다.
    그래서 '덮개가 걷혔다' 와 '페이지가 앉았다' 둘 다 기다린다."""
    from aoi_verification.app.ui import main_window as mw
    from aoi_verification.app.ui.pages import match_review_page as mrp

    nxt = inspect.getsource(mrp.MatchReviewPage._make_next_rows)
    assert "hide_overlay(then=self._arm_entrance)" in nxt, (
        "덮개가 걷히기 전에 재생한다")
    maybe = inspect.getsource(mrp.MatchReviewPage._maybe_play_entrance)
    assert "isVisible()" in maybe, "화면 밖에서도 재생한다"
    assert hasattr(mrp.MatchReviewPage, "on_shown")
    # 창이 '앉았다' 를 알려 주는 쪽도 있어야 짝이 맞는다.
    assert "_notify_shown" in inspect.getsource(mw.MainWindow._show_page)


def test_swipe_touches_only_the_snapshot_never_the_live_photo():
    """모션은 **떠나는 그림의 사본에만** 건다(시안 명시).

    ★ 살아 있는 `center_img` 에 걸면 두 가지가 깨진다: (a) `_advance_incremental`
    이 곧바로 같은 위젯에 다음 사진을 넣으므로 **새 사진이** 밀려나고, (b) 그
    위젯에 그래픽스 이펙트가 붙어 있는 동안 사진이 다시 그려질 때마다 오프스크린
    렌더가 따라붙는다."""
    from aoi_verification.app.ui.pages import select_page as sp

    src = inspect.getsource(sp.SelectPage._swipe_out_decided)
    assert "img.grab()" in src, "사본을 뜨지 않는다"
    assert 'QPropertyAnimation(ghost, b"pos"' in src, "이동이 pos 애니가 아니다"
    assert "center_img)" not in src.replace("img = self.center_img", ""), (
        "살아 있는 사진 위젯에 모션을 걸고 있다")
    # 들어오는 사진은 **앞 사진이 다 빠진 뒤** 들어온다(시안 CSS 의 `.18s backwards`).
    fade = inspect.getsource(sp.SelectPage._fade_in_next)
    assert "delay_ms=motion.DUR_SWIPE_OUT" in fade, (
        "겹쳐 재생한다 — 한순간 화면이 비어 보인다")


def test_rapid_decisions_skip_the_animation(styled_qapp, motion_on, tmp_path):
    """연타 중에는 **생략한다**(시안 명시) — 없으면 애니가 쌓여 버벅인다.

    실측(900px 사진 39회 연타, 결정당 평균): 모션 OFF 28.3ms · 가드 없음
    33.7ms(최악 78.7) · 가드 있음 28.2ms(최악 49.8).  가드가 있으면 공짜다.
    ★ 앞 잔상은 **즉시 치운다** — 남겨 두면 지지난 사진이 지금 사진 위에 뜬다."""
    from PyQt6.QtGui import QColor, QImage

    from aoi_verification.app.models.slot import ImageItem
    from aoi_verification.app.ui.pages import select_page as sp

    for i in range(3):
        im = QImage(120, 120, QImage.Format.Format_RGB32)
        im.fill(QColor(200 - 40 * i, 90, 120))
        im.save(str(tmp_path / f"p{i}.jpg"), "JPG", 90)
    page = sp.SelectPage()
    page.resize(900, 700)
    page.show()
    page.load_state(
        queue=[ImageItem(slot="S1", path=tmp_path / f"p{i}.jpg", side="ref")
               for i in range(3)],
        targets={}, excluded={}, history=[])
    for _ in range(8):
        styled_qapp.processEvents()

    page._decide("verify")
    assert getattr(page, "_swipe_ghost", None) is not None, "첫 결정은 움직인다"
    page._decide("exclude")               # 앞 애니가 도는 중 = 연타
    assert getattr(page, "_swipe_ghost", None) is None, (
        "연타인데 새 애니를 걸었다 — 결정마다 쌓인다")
    page.close()


# ═══ 28안 — 출력 엑셀 양식 ═══════════════════════════════════════════════════
def _tiny_result(tmp: Path):
    from aoi_verification.app.models.result import (FinalResult, MatchResult,
                                                    MissEntry)
    return FinalResult(
        mode="single", ref_machine="1", val_machine="2",
        matches=[MatchResult(slot="S1", ref_path=tmp / "a.jpg",
                             val_path=tmp / "b.jpg", score=0.9)],
        slot_only_ref=[], slot_only_val=[],
        unmatched_refs=[MissEntry(slot="S1", side="ref", path=tmp / "c.jpg",
                                  note="x"),
                        MissEntry(slot="S2", side="ref", path=tmp / "d.jpg",
                                  note="x")])


@pytest.fixture
def exported(isolated_cache):
    from openpyxl import load_workbook

    from aoi_verification.app.workers.exporter import ExcelExporter

    tmp = Path(tempfile.mkdtemp())
    dst = tmp / "out.xlsx"
    exp = ExcelExporter(_tiny_result(tmp), dst,
                        template_path=tmp / "none.xlsx",
                        include_full_template=True)
    msgs: list[str] = []
    exp.signals.progress.connect(lambda d, t, m: msgs.append(m))
    exp.run()
    return load_workbook(dst), dst, msgs


def test_unmatched_rows_are_tinted(exported):
    """미매칭 행은 A~D 가 통째로 물든다 — 흑백 인쇄에서도 세어진다.

    ★ 요약 시트는 매치와 미매칭이 **한 표에 섞여** 정렬된다.  표시가 D열 빨간
    글씨뿐이면 '총 몇 건 중 몇 건이 미매칭인지' 를 한 줄씩 훑어야 하고, 흑백으로
    인쇄하면 그 빨강마저 사라진다."""
    from aoi_verification.app.workers import exporter as ex

    wb, dst, _ = exported
    ws = wb[dst.stem]
    tinted = [r for r in (3, 4, 5)
              if str(ws[f"A{r}"].fill.fgColor.rgb).upper() == ex.UNMATCHED_FILL]
    assert len(tinted) == 2, "미매칭 2건이 물들지 않았다"
    assert len(tinted) < 3, "매치 행까지 물들었다 — 구분이 사라진다"


def test_every_sheet_freezes_and_repeats_the_header(exported):
    """머리 2행은 표의 이름표다 — 스크롤·인쇄 어느 쪽에서도 살아 있어야 한다.

    ★ `wb.create_sheet` 로 만든 시트는 양식의 설정을 **하나도** 물려받지 않는다."""
    wb, _dst, _ = exported
    for name in wb.sheetnames:
        ws = wb[name]
        assert ws.freeze_panes == "A3", f"{name}: 머리 행이 안 얼었다"
        # openpyxl 이 절대 참조로 정규화한다($1:$2) — 행 범위만 비교한다.
        titles = str(ws.print_title_rows or "").replace("$", "")
        assert titles == "1:2", f"{name}: 인쇄 반복 헤더가 없다({titles!r})"


# ═══ 30안 — 저장 진행률: 어느 시트의 어느 슬롯인지 ═══════════════════════════
def test_progress_says_which_sheet_and_slot(exported):
    """진행 문구는 시트와 슬롯을 말한다.  행 수치는 진행 라벨의 몫이다(단일 출처)."""
    _wb, dst, msgs = exported
    assert msgs, "진행 보고가 없다"
    named = set(msgs)
    assert any(i18n.KO.SHEET_UNMATCHED in m and "S1" in m for m in named), named
    assert any(dst.stem in m for m in named), named
    for m in named:
        assert "/" not in m, f"문구에 수치가 들어갔다: {m!r}"


# ═══ 29안 · 31안 — 내보내기 옵션 · 저장 목적지 ═══════════════════════════════
@pytest.fixture
def page(styled_qapp):
    from aoi_verification.app.ui.pages import result_page as rp

    p = rp.ResultPage()
    p.resize(1512, 950)
    p.show()
    for _ in range(6):
        styled_qapp.processEvents()
    yield p
    p.close()


def test_subsumed_option_is_disabled_but_keeps_its_value(page, styled_qapp):
    """'사진을 원본 화질로' 는 '미매칭만 원본' 을 삼킨다 — 화면이 그 사실을 말한다.

    ★ 체크 상태는 건드리지 않는다.  전역 옵션을 껐을 때 원래 두었던 선택이 그대로
    돌아와야 한다 — 끄는 순간 값이 바뀌면 '내가 언제 저걸 껐지' 가 된다."""
    page.unmatched_original_chk.setChecked(True)
    page.original_quality_chk.setChecked(True)
    styled_qapp.processEvents()
    assert not page.unmatched_original_chk.isEnabled()
    assert page.unmatched_original_chk.isChecked(), "값을 멋대로 바꿨다"
    # 툴팁은 그대로 유지한다(시안 명시).
    assert page.unmatched_original_chk.toolTip() == (
        i18n.KO.EXPORT_UNMATCHED_ORIGINAL_TOOLTIP)

    page.original_quality_chk.setChecked(False)
    styled_qapp.processEvents()
    assert page.unmatched_original_chk.isEnabled()


def test_option_row_and_order_are_unchanged(page):
    """29안은 **표현만** 바꾼다 — 배치·순서·기본값은 현행 그대로."""
    def left(w):
        return w.mapTo(page, w.rect().topLeft()).x()

    assert left(page.unmatched_original_chk) < left(page.original_quality_chk)
    assert left(page.original_quality_chk) < left(page.full_template_chk)
    for chk in (page.unmatched_original_chk, page.original_quality_chk,
                page.full_template_chk):
        assert not chk.isChecked(), "기본값이 바뀌었다"


def test_save_destination_is_shown_before_and_after_saving(page, styled_qapp):
    """저장 전에는 '무엇이 생기는가', 저장 뒤에는 '같은 파일에 덮어쓴다'.

    ★ 두 번째 저장이 대화상자 없이 조용히 덮어쓰던 것이 이 화면에서 가장
    놀라운 동작이었다 — 예고가 되면 놀람이 아니다."""
    tmp = Path(tempfile.mkdtemp())
    target = tmp / "AOI 18호기 검증 (17호기 기준).xlsx"
    page.show_result(_tiny_result(tmp), target_path=target)
    styled_qapp.processEvents()
    assert target.name in page.save_target_label.text()
    assert page.save_target_label.toolTip() == str(target), "전체 경로는 툴팁이"

    page._exported = True
    page._save_path = target
    page._refresh_save_target()
    text = page.save_target_label.text()
    assert "덮어" in text, f"재저장이 덮어쓰기임을 말하지 않는다: {text!r}"


