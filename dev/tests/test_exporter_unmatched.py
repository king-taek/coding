"""Excel exporter — 매칭 + 미매칭 reference 통합 출력 검증."""

import os
import tempfile
from pathlib import Path

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
pytest.importorskip("PyQt6.QtWidgets")
pytest.importorskip("openpyxl")
pytest.importorskip("PIL.Image")

from PIL import Image                                          # noqa: E402
from PyQt6.QtWidgets import QApplication                       # noqa: E402

from aoi_verification.app.models.result import (               # noqa: E402
    FinalResult, MatchResult, MissEntry,
)
from aoi_verification.app.workers.exporter import ExcelExporter  # noqa: E402


@pytest.fixture(scope="module")
def qapp():
    return QApplication.instance() or QApplication([])


def _make_image(folder: Path, name: str) -> Path:
    p = folder / name
    Image.new("RGB", (400, 300), color=(80, 120, 200)).save(str(p), "JPEG")
    return p


def test_export_writes_matches_and_unmatched(qapp, isolated_cache, tmp_path):
    """매칭 + 미매칭이 같이 들어가면 결과 워크북에 두 종류 모두 나타난다."""
    src = tmp_path / "src"
    src.mkdir()
    a_ref = _make_image(src, "a_ref.jpeg")
    a_val = _make_image(src, "a_val.jpeg")
    b_ref = _make_image(src, "b_ref.jpeg")  # 미매칭

    result = FinalResult(
        mode="single",
        ref_machine="1호기",
        val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=a_ref, val_path=a_val,
                              score=0.9)],
        unmatched_refs=[MissEntry(slot="S1", side="ref", path=b_ref,
                                   note="미매칭")],
    )
    dst = tmp_path / "out.xlsx"
    # 양식 자동 감지를 피하려고 일부러 존재하지 않는 템플릿 경로를 명시 지정.
    no_tpl = tmp_path / "no_template.xlsx"
    exp = ExcelExporter(result, dst_path=dst, template_path=no_tpl)
    exp.run()  # QThread.run() 을 동기 실행
    assert dst.exists()

    from aoi_verification.app import i18n
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import CellRichText
    wb = load_workbook(str(dst), rich_text=True)
    # 요약 시트(파일명 stem)를 이름으로 직접 읽는다 — .active 는 openpyxl 버전에
    # 따라 미매칭 시트를 가리킬 수 있어 의존하지 않는다.
    ws = wb["out"]
    # 헤더는 row 1~2, 데이터는 row 3 부터. 정렬: a_ref(매칭)=row3, b_ref(미매칭)=row4.
    assert ws["D3"].value is None, "매칭 행의 D 셀은 비어있어야 (이미지만 임베드)"
    d4 = ws["D4"].value
    # geometry 기능 활성 + Surface.flt 없음 → 파일명 + '미지원 자재' 마커(rich-text).
    assert "b_ref.jpeg" in str(d4)
    assert i18n.KO.GEOM_NOT_SUPPORTED in str(d4)
    # 파일명 블록은 여전히 빨강.
    assert isinstance(d4, CellRichText)
    assert "FF2D55" in str(d4[0].font.color.rgb).upper()

    # 코멘트 ‘미매칭’ 확인
    assert ws["D4"].comment is not None
    assert "미매칭" in str(ws["D4"].comment.text)

    # AOI-N 헤더가 row 2 에 들어갔는지 (#3).
    assert ws["C2"].value == "AOI-1"
    assert ws["D2"].value == "AOI-2"


def test_export_no_unmatched_unaffected(qapp, isolated_cache, tmp_path):
    """unmatched_refs 가 비어있으면 기존 동작과 동일하게 매칭만 출력."""
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "ref.jpeg")
    val = _make_image(src, "val.jpeg")

    result = FinalResult(
        mode="single",
        ref_machine="1호기",
        val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val,
                              score=0.9)],
    )
    dst = tmp_path / "out2.xlsx"
    no_tpl = tmp_path / "no_template.xlsx"
    ExcelExporter(result, dst_path=dst, template_path=no_tpl).run()
    assert dst.exists()

    from openpyxl import load_workbook
    ws = load_workbook(str(dst)).active
    # 헤더 2 줄 + 데이터 row 3 부터. 행 3 = 매칭, 행 4 는 데이터 없음.
    assert ws["B3"].value == "S1"
    assert ws["D3"].value is None  # 이미지 임베드만, 텍스트 없음
    assert ws["B4"].value is None


def test_export_full_template_optional(qapp, isolated_cache, tmp_path):
    """전체 양식(E~H 포함) 시트는 옵션 — 기본(off)이면 생성 안 하고, 켜면 생성(#3)."""
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "ref.jpeg")
    val = _make_image(src, "val.jpeg")
    result = FinalResult(
        mode="single", ref_machine="1호기", val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val, score=0.9)],
    )
    from aoi_verification.app.utils import paths
    from openpyxl import load_workbook

    # 기본(off) — 요약 시트만, '전체 양식' 없음.
    dst = tmp_path / "결과_AOI-1_vs_AOI-2.xlsx"
    ExcelExporter(result, dst_path=dst,
                  template_path=paths.template_path()).run()
    wb = load_workbook(str(dst))
    assert wb.sheetnames[0] == "결과_AOI-1_vs_AOI-2"
    assert "전체 양식" not in wb.sheetnames
    summary = wb["결과_AOI-1_vs_AOI-2"]
    # 요약 시트는 A~D 만 — E~H 헤더 없음 + 데이터 채워짐.
    assert summary["A1"].value == "No" and summary["C1"].value == "Scan Defect"
    assert summary["E1"].value is None and summary["G1"].value is None
    assert summary["B3"].value == "S1"

    # 옵션 on — '전체 양식'(E~H 포함) 시트도 생성.
    dst2 = tmp_path / "결과2.xlsx"
    ExcelExporter(result, dst_path=dst2, template_path=paths.template_path(),
                  include_full_template=True).run()
    wb2 = load_workbook(str(dst2))
    assert "전체 양식" in wb2.sheetnames
    assert wb2.sheetnames.index("전체 양식") > 0          # 요약이 앞, 전체가 뒤
    full = wb2["전체 양식"]
    assert full["E1"].value is not None and full["G1"].value is not None
    assert full["B3"].value == "S1"


def test_export_unmatched_sheet(qapp, isolated_cache, tmp_path):
    """미매칭이 있으면 '미매칭 사진' 시트가 추가된다(이미지 포함, #3)."""
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "ref.jpeg")
    val = _make_image(src, "val.jpeg")
    miss = _make_image(src, "miss.jpeg")
    result = FinalResult(
        mode="single", ref_machine="1호기", val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val, score=0.9)],
        unmatched_refs=[MissEntry(slot="S2", side="ref", path=miss, note="미매칭")],
    )
    from aoi_verification.app import i18n
    from openpyxl import load_workbook
    no_tpl = tmp_path / "no_template.xlsx"
    dst = tmp_path / "out.xlsx"
    ExcelExporter(result, dst_path=dst, template_path=no_tpl).run()
    wb = load_workbook(str(dst), rich_text=True)
    assert i18n.KO.SHEET_UNMATCHED in wb.sheetnames
    um = wb[i18n.KO.SHEET_UNMATCHED]
    # 미매칭 시트엔 미매칭 행만 — 첫 데이터 행 slot=S2, D열에 파일명(+미지원 마커).
    assert um["B3"].value == "S2"
    assert "miss.jpeg" in str(um["D3"].value)
    assert i18n.KO.GEOM_NOT_SUPPORTED in str(um["D3"].value)


def _install_flt_schema(monkeypatch):
    """exporter 테스트용 — surface_flt 에 합성 스키마(32B) 설치(zone/recipe 포함)."""
    from aoi_verification.app.coords import surface_flt, camtek_ini
    fields = {"actual_x": (0, "f"), "actual_y": (4, "f"), "area": (8, "f"),
              "blob_breadth": (12, "f"), "blob_feret_max": (16, "f"),
              "contrast": (20, "f"), "zone": (24, "B"), "recipe": (25, "B")}
    monkeypatch.setattr(surface_flt, "_FIELDS", dict(fields))
    monkeypatch.setattr(surface_flt, "_RECORD_SIZE", 32)
    monkeypatch.setattr(surface_flt, "_HEADER_BYTES", 0)
    monkeypatch.setattr(surface_flt, "_SCHEMA_READY", True)
    surface_flt.load_folder.cache_clear()
    camtek_ini.load_abs_folder.cache_clear()


def _write_flt_record(folder: Path, x, y, area, breadth, feret, contrast,
                      zone=1, recipe=2):
    import struct
    buf = bytearray(32)
    for off, v in ((0, x), (4, y), (8, area), (12, breadth),
                   (16, feret), (20, contrast)):
        struct.pack_into("<f", buf, off, float(v))
    struct.pack_into("<B", buf, 24, int(zone))
    struct.pack_into("<B", buf, 25, int(recipe))
    (folder / "Surface.flt").write_bytes(bytes(buf))


def test_unmatched_geometry_rendered(qapp, isolated_cache, tmp_path, monkeypatch):
    """Surface.flt + 좌표 일치 → 미매칭 D셀에 geometry(area/contrast) 표기."""
    _install_flt_schema(monkeypatch)
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "a_ref.jpeg")
    val = _make_image(src, "a_val.jpeg")
    miss = _make_image(src, "z_miss.jpeg")
    _write_flt_record(src, 1000.0, 2000.0, 55.0, 2.0, 11.0, 108.0)
    (src / "ColorImageGrabingInfo.ini").write_text(
        "[z_miss.jpeg]\nX=1000.0\nY=2000.0\nCol=3\nRow=5\n", encoding="utf-8")

    result = FinalResult(
        mode="single", ref_machine="1호기", val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val, score=0.9)],
        unmatched_refs=[MissEntry(slot="S1", side="ref", path=miss, note="미매칭")],
    )
    dst = tmp_path / "out.xlsx"
    no_tpl = tmp_path / "no_template.xlsx"
    ExcelExporter(result, dst_path=dst, template_path=no_tpl).run()

    from aoi_verification.app import i18n
    from openpyxl import load_workbook
    # 미매칭 사진 시트에는 미매칭 행만 — z_miss 가 첫 데이터 행(row3).
    ws = load_workbook(str(dst), rich_text=True)[i18n.KO.SHEET_UNMATCHED]
    d3 = str(ws["D3"].value)
    assert "z_miss.jpeg" in d3
    assert "area" in d3 and "contrast" in d3
    assert "zone" in d3 and "recipe" in d3
    # contrast 108 (비0) → 값이 그대로, '—' 아님.
    assert "108" in d3 and "contrast —" not in d3


def test_unmatched_geometry_contrast_zero_dash(qapp, isolated_cache, tmp_path,
                                               monkeypatch):
    """contrast=0(대부분 자재) → 'contrast —' 로 표기(0.00 아님)."""
    _install_flt_schema(monkeypatch)
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "a_ref.jpeg")
    val = _make_image(src, "a_val.jpeg")
    miss = _make_image(src, "z_miss.jpeg")
    _write_flt_record(src, 1000.0, 2000.0, 55.0, 2.0, 11.0, 0.0, zone=63)
    (src / "ColorImageGrabingInfo.ini").write_text(
        "[z_miss.jpeg]\nX=1000.0\nY=2000.0\nCol=3\nRow=5\n", encoding="utf-8")
    result = FinalResult(
        mode="single", ref_machine="1호기", val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val, score=0.9)],
        unmatched_refs=[MissEntry(slot="S1", side="ref", path=miss, note="미매칭")],
    )
    dst = tmp_path / "out.xlsx"
    ExcelExporter(result, dst_path=dst,
                  template_path=tmp_path / "no_template.xlsx").run()
    from aoi_verification.app import i18n
    from openpyxl import load_workbook
    d3 = str(load_workbook(str(dst), rich_text=True)[i18n.KO.SHEET_UNMATCHED]["D3"].value)
    assert "contrast —" in d3 and "contrast 0.00" not in d3


def test_unmatched_no_flt_marker(qapp, isolated_cache, tmp_path, monkeypatch):
    """스키마는 켜졌지만 Surface.flt 가 없으면 '미지원 자재' 마커가 붙는다."""
    from aoi_verification.app import i18n
    _install_flt_schema(monkeypatch)
    src = tmp_path / "src"
    src.mkdir()
    ref = _make_image(src, "a_ref.jpeg")
    val = _make_image(src, "a_val.jpeg")
    miss = _make_image(src, "z_miss.jpeg")  # Surface.flt 없음

    result = FinalResult(
        mode="single", ref_machine="1호기", val_machine="2호기",
        matches=[MatchResult(slot="S1", ref_path=ref, val_path=val, score=0.9)],
        unmatched_refs=[MissEntry(slot="S1", side="ref", path=miss, note="미매칭")],
    )
    dst = tmp_path / "out.xlsx"
    no_tpl = tmp_path / "no_template.xlsx"
    ExcelExporter(result, dst_path=dst, template_path=no_tpl).run()

    from openpyxl import load_workbook
    ws = load_workbook(str(dst), rich_text=True)[i18n.KO.SHEET_UNMATCHED]
    d3 = str(ws["D3"].value)
    assert "z_miss.jpeg" in d3
    assert i18n.KO.GEOM_NOT_SUPPORTED in d3


def test_machine_label_rule():
    """호기 라벨 규칙: 숫자/N호기 → AOI-N, 그 외 문자 포함 → AOI(원본)."""
    from aoi_verification.app.workers.exporter import _machine_label
    # 순수 숫자 / N호기 → AOI-N
    assert _machine_label("2") == "AOI-2"
    assert _machine_label("2호기") == "AOI-2"
    assert _machine_label("1호기") == "AOI-1"
    assert _machine_label(" 7 호기 ") == "AOI-7"
    assert _machine_label("10") == "AOI-10"
    # 다른 문자 포함 → AOI(원본값)
    assert _machine_label("K-2") == "AOI(K-2)"
    assert _machine_label("K-6") == "AOI(K-6)"
    assert _machine_label("AOI-3") == "AOI(AOI-3)"
    # 빈 입력 → ""
    assert _machine_label("") == ""
    assert _machine_label("   ") == ""
    assert _machine_label(None) == ""
