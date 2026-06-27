"""flt_check_auto_discover — 합성 NAS 트리로 탐색·파싱·매칭·분류 검증.

탐색/파싱은 stdlib 만 → 헤드리스 통과. Excel 생성은 openpyxl 게이트.
"""

import importlib.util
import struct
from pathlib import Path

import pytest

# dev/flt_check_auto_discover.py 를 모듈로 로드(패키지 아님).
_MOD_PATH = Path(__file__).resolve().parents[1] / "flt_check_auto_discover.py"
_spec = importlib.util.spec_from_file_location("flt_check_auto_discover", _MOD_PATH)
ad = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(ad)


# 실측 픽스처 record 2건(zone1 contrast≠0, zone63 contrast0).
_REAL = [
    "00001d0000000020906dab41602b134138fb1269788e0e41906dab41602b134138fb1269788e0e41010000006900000000000000000000000000000000010200000000000000144000000000000018400000000000003840000000e089dc09400000008059b21d40000000c0ff8919400000000000e0504000000000000014400000000000000000000000a0aaea5e400000000000000000",
    "1d00000000040000a6756ba85096134148d30a475e9b0f41a6756ba85096134148d30a475e9b0f41ffffffff0f0000083613000000a0404400a01344003f0200000000000000284000000000000024400000000000804740000000c00f3f0a4000000080d0a62c40000000c0d3ae2b400000000000803640000000c076f81440000000000000000000000000000000000000000000000000",
]
# (actual_x, actual_y) of each record — INI 좌표를 여기에 맞춘다.
_AX = [(314072.064131, 250319.051306), (320916.164472, 258923.784689)]


def _build_tree(tmp_path: Path) -> Path:
    """root/R_TB500_x/Setup1/LOT1/wafer/{Surface.flt,INI} 구조 생성."""
    wafer = tmp_path / "R_TB500_LIVE_x" / "Setup1" / "LOT1" / "waferA"
    wafer.mkdir(parents=True)
    (wafer / "Surface.flt").write_bytes(b"".join(bytes.fromhex(h) for h in _REAL))
    # 두 record 좌표에 맞춘 INI 두 섹션 + 매칭 안 되는 먼 좌표 1건(aux 후보).
    ini = (f"[d_zone1.jpeg]\nFaultX={_AX[0][0]}\nFaultY={_AX[0][1]}\nCol=3\nRow=5\nRecipeNumber=2\n"
           f"[d_zone63.jpeg]\nFaultX={_AX[1][0]}\nFaultY={_AX[1][1]}\nCol=6\nRow=5\nRecipeNumber=2\n"
           f"[d_far.jpeg]\nFaultX=999999\nFaultY=999999\nCol=1\nRow=1\nRecipeNumber=2\n")
    (wafer / "ColorImageGrabingInfo.ini").write_text(ini, encoding="utf-8")
    return wafer


def test_parse_flt_real_records(tmp_path):
    p = tmp_path / "Surface.flt"
    p.write_bytes(b"".join(bytes.fromhex(h) for h in _REAL))
    recs, size, framed = ad.parse_flt(p)
    assert framed and size % 152 == 0 and len(recs) == 2
    assert recs[0]["zone"] == 1 and recs[0]["recipe"] == 2
    assert round(recs[0]["contrast"], 1) == 123.7
    assert recs[1]["zone"] == 63 and recs[1]["contrast"] == 0


def test_discover_finds_tb500_wafer(tmp_path):
    wafer = _build_tree(tmp_path)
    found = ad.discover(tmp_path, tb500_only=True, latest_lots=3, max_depth=5,
                        per_root_max=20, deep=False, deadline=ad.time.time() + 30,
                        log=lambda m: None)
    assert wafer in found


def test_discover_skips_non_tb500(tmp_path):
    # TB500 아닌 제품 폴더만 있으면(엄격 모드) 못 찾는다.
    w = tmp_path / "SomeOther_Product" / "Setup1" / "LOT1" / "waferA"
    w.mkdir(parents=True)
    (w / "Surface.flt").write_bytes(bytes.fromhex(_REAL[0]))
    found = ad.discover(tmp_path, tb500_only=True, latest_lots=3, max_depth=5,
                        per_root_max=20, deep=False, deadline=ad.time.time() + 30,
                        log=lambda m: None)
    assert found == []
    # --all 동등(tb500_only=False) 이면 찾는다.
    found2 = ad.discover(tmp_path, tb500_only=False, latest_lots=3, max_depth=5,
                         per_root_max=20, deep=False, deadline=ad.time.time() + 30,
                         log=lambda m: None)
    assert w in found2


def test_process_matches_and_classifies(tmp_path):
    wafer = _build_tree(tmp_path)
    rows, schema = [], []
    ad.process_folder(wafer, tol=5.0, rows=rows, schema_rows=schema,
                      log=lambda m: None)
    assert schema[0]["framed_ok"] and schema[0]["records"] == 2
    assert schema[0]["ini_entries"] == 3
    matched = [r for r in rows if r["record_index"] != ""]
    failed = [r for r in rows if r["record_index"] == ""]
    assert len(matched) == 2 and len(failed) == 1
    # 매칭된 zone1 행 — recipe_match Y, contrast≠0, UI 열은 빈칸.
    z1 = next(r for r in matched if str(r["zone"]) == "1")
    assert z1["recipe_match"] == "Y"
    assert float(z1["contrast"]) > 0
    assert z1["ui_contrast"] == "" and z1["ui_match"] == ""
    # zone63 행 contrast==0.
    z63 = next(r for r in matched if str(r["zone"]) == "63")
    assert float(z63["contrast"]) == 0
    # 실패 1건은 aux/revisit 후보(ini>records) 로 분류.
    assert "aux" in failed[0]["note"]


def test_crosstab_and_shortlist(tmp_path):
    wafer = _build_tree(tmp_path)
    rows, schema = [], []
    ad.process_folder(wafer, 5.0, rows, schema, lambda m: None)
    ct = ad.zone_recipe_crosstab(rows)
    # (zone1,recipe2) 와 (zone63,recipe2) 두 버킷.
    keys = {(c["zone"], c["recipe"]) for c in ct}
    assert (1, 2) in keys and (63, 2) in keys
    sl = ad.ui_shortlist(rows)
    assert any(s["추출_zone"] == 1 for s in sl)
    assert all(s["UI_contrast"] == "" for s in sl)  # UI 칸은 빈칸


def test_excel_output(tmp_path):
    pytest.importorskip("openpyxl")
    wafer = _build_tree(tmp_path)
    rows, schema = [], []
    ad.process_folder(wafer, 5.0, rows, schema, lambda m: None)
    out = tmp_path / "out.xlsx"
    ad.write_excel(str(out), [wafer], schema, rows, ["log1"])
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    for name in ["자동탐색_대상폴더", "폴더_스키마_정합성", "검증결과_전체",
                 "zone별_contrast", "UI수기확인_shortlist", "요약_결론", "실행로그"]:
        assert name in wb.sheetnames
