"""엑셀(`양식.xlsx`) 출력 워커.

- 양식.xlsx 를 그대로 템플릿으로 로드해서 셀 서식/병합/열폭/행높이를 보존한다.
- 헤더(C2/D2)의 ‘AOI-N’ 만 검증 세션의 호기 번호로 교체.
- 데이터: A=번호, B=Slot, C=기준(낮은 호기) 이미지, D=검증(높은 호기) 이미지.
- E~H 컬럼(Escape Defect Camtek/KLA)은 사용자가 수기로 채울 영역이라 비워 둔다.
- ‘매칭 방향’ 컬럼은 사용자 요청으로 더 이상 쓰지 않는다 (#4).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

from PyQt6.QtCore import QObject, QThread, pyqtSignal

from .. import i18n
from ..models.result import FinalResult, MatchResult, MissEntry
from ..utils import image_io


# ---------------------------------------------------------------------------
# 양식.xlsx 의 컬럼 레이아웃 (사용자 양식과 1:1)
#   A = No / B = slot# / C = 기준(낮은 호기) / D = 검증(높은 호기)
# Header 는 두 줄: row 1 (그룹) + row 2 (호기 번호 = ‘AOI-N’). 데이터는 row 3 부터.
# ---------------------------------------------------------------------------
COL_NO = "A"
COL_SLOT = "B"
COL_REF = "C"
COL_VAL = "D"
DATA_START_ROW = 3
HEADER_AOI_ROW = 2
# 시트 분리 (#사용자 요청): 1번째=요약(A~D, 파일명), 2번째=전체 양식(E~H 포함).
SHEET_FULL_NAME = "전체 양식"
# 슬롯 구분선이 그려지는 컬럼 — 사용자 요청 (#6): A~D 만, 두껍게.
BORDER_COLS = ["A", "B", "C", "D"]

# 셀 ↔ 사진 크기 정합:
#   · 양식.xlsx 의 데이터 행 높이 (165.75pt) 와 일치시켜 템플릿 안팎의 행 높이를
#     맞춘다. 양식이 없을 때(폴백) 도 동일 값으로 통일.
#   · 이미지 max 변 = 150 px. 1pt ≈ 1.333 px 이므로 165pt ≈ 220 px → 150 px
#     이미지가 셀 안에 여유 있게 들어간다.
ROW_HEIGHT_PT = 165.75
IMG_COL_WIDTH = 22


def _machine_label(raw: str) -> str:
    """호기 입력을 엑셀 헤더 라벨로 정규화.

    - 순수 숫자(``2``) 또는 ``N호기``(``2호기``, 공백 허용) → ``AOI-N``
    - 그 외 문자가 포함되면(``K-2`` 등) → ``AOI(원본값)``
    - 빈 입력 → ``""``
    """
    s = (raw or "").strip()
    if not s:
        return ""
    m = re.fullmatch(r"(\d+)(\s*호기)?", s)
    if m:
        return f"AOI-{m.group(1)}"
    return f"AOI({s})"


class ExporterSignals(QObject):
    progress = pyqtSignal(int, int, str)
    done = pyqtSignal(str)               # 결과 파일 경로
    failed = pyqtSignal(str)


class ExcelExporter(QThread):
    """`FinalResult` 를 받아 양식.xlsx 템플릿에 채워 저장."""

    def __init__(self,
                 result: FinalResult,
                 dst_path: Path,
                 template_path: Optional[Path] = None,
                 include_full_template: bool = False,
                 original_quality: bool = False,
                 parent: QObject | None = None) -> None:
        super().__init__(parent)
        self._result = result
        self._dst = Path(dst_path)
        self._template = Path(template_path) if template_path else None
        # 전체 양식(E~H 수기 영역 포함) 시트 생성 여부 — 기본 off(가볍고 빠른 출력).
        self._include_full_template = bool(include_full_template)
        # 사진을 원본 화질로 임베드할지 — 기본 off(중간 화질 캐시로 가볍게).
        self._original_quality = bool(original_quality)
        self.signals = ExporterSignals()

    # ------------------------------------------------------------------
    def run(self) -> None:        # type: ignore[override]
        try:
            self._do_export()
            self.signals.done.emit(str(self._dst))
        except Exception as exc:
            self.signals.failed.emit(str(exc))

    # ------------------------------------------------------------------
    def _do_export(self) -> None:
        from openpyxl import Workbook, load_workbook

        # 양식이 있으면 그대로 로드해서 셀 서식을 모두 보존.
        # 없으면 동일 컬럼 구조의 워크북을 빈 상태로 만든다.
        if self._template is not None and self._template.exists():
            wb = load_workbook(str(self._template))
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "AOI 검증 결과"
            self._build_minimal_headers(ws)

        # 시트는 둘로 나눈다 (#사용자 요청):
        #   · 1번째 시트 = 결과 파일명과 같은 이름, A~D 열만(요약).
        #   · 2번째 시트 = 기존 양식 그대로(전체 — E~H 수기 영역 포함), 이름 '전체 양식'.
        # 구현: 템플릿(현재 ws)을 채워 '전체 양식' 으로 두고, 그 시트를 복제해 E~H 를
        #       지운 요약 시트를 앞쪽에 만든다(이미지/서식 보존을 위해 채운 뒤 복제).
        ws.title = SHEET_FULL_NAME

        # row 2 의 ‘AOI-N’ 헤더를 실제 호기 번호로 교체 (#3).
        ref_label = _machine_label(self._result.ref_machine)
        val_label = _machine_label(self._result.val_machine)
        if ref_label:
            ws[f"{COL_REF}{HEADER_AOI_ROW}"] = ref_label
        if val_label:
            ws[f"{COL_VAL}{HEADER_AOI_ROW}"] = val_label

        # 컬럼 폭 보정 — 양식.xlsx 는 ‘Scan Defect (C1:D1)’ 같은 병합 헤더의
        # 왼쪽 셀에만 width 를 지정해 두어, 오른쪽 셀(D, F, H 등) 이 기본 폭
        # (~8) 으로 떨어져 사진이 작아 보이는 문제가 있다.  병합된 헤더 쌍의
        # 왼쪽 컬럼 width 를 오른쪽 컬럼에도 그대로 미러링한다.
        self._mirror_paired_column_widths(ws)
        # C, D, E, F, G, H 모두 같은 폭 — 양식 C(31.83) 를 기준값으로 통일
        # (#3 — 사용자 요청: D 도 C 와 같고, E~H 도 모두 동일).  사진이 들어가는
        # C/D 와 사용자 수기 영역인 E~H 가 시각적으로 일관된 폭을 갖도록.
        self._equalize_column_group(
            ws, [COL_REF, COL_VAL, "E", "F", "G", "H"],
            floor=IMG_COL_WIDTH,
        )
        self._ensure_width(ws, COL_SLOT, 14)
        self._ensure_width(ws, COL_NO, 6)

        # 매칭/미매칭 통합 정렬 → Slot 오름차순, 그 안에서 기준 파일명 오름차순.
        rows_input: list[tuple[str, str, object]] = []
        for m in self._result.matches:
            rows_input.append((m.slot, str(m.ref_path.name).lower(), m))
        for u in self._result.unmatched_refs:
            rows_input.append((u.slot, str(u.path.name).lower(), u))
        rows_input.sort(key=lambda x: (x[0], x[1]))

        # 전체 양식(E~H 포함) 시트는 옵션 — 기본 off 면 이미지 임베드를 1회만 하게
        # 요약 시트만 채운다(더 빠르고 가벼운 파일).  켜면 전체 양식도 채운다.
        if self._include_full_template:
            self._fill_rows(ws, rows_input)
            # 양식.xlsx 의 A3..A22 미리 박힌 1..20 행번호 중 안 채운 행은 비운다.
            data_end_row = DATA_START_ROW + len(rows_input) - 1
            for r in range(max(data_end_row + 1, DATA_START_ROW), ws.max_row + 1):
                a = ws.cell(row=r, column=1)
                if isinstance(a.value, (int, float)):
                    a.value = None

        # 시트 순서: 미매칭(첫 번째, 조건부) → 요약 → 전체 양식.
        unmatched_rows = [r for r in rows_input if isinstance(r[2], MissEntry)]
        if unmatched_rows:
            # 미매칭 시트를 index 0(첫 번째)에 만들고, 요약은 index 1.
            self._write_unmatched_sheet(wb, unmatched_rows)
            self._build_summary_sheet(wb, rows_input, index=1)
        else:
            self._build_summary_sheet(wb, rows_input, index=0)

        # Slot 불일치 ---------------------------------------------------
        if self._result.slot_only_ref or self._result.slot_only_val:
            self._write_slot_mismatch_sheet(wb)

        # 전체 양식 미포함이면, 헤더 복사가 끝난 지금 전체 양식 시트를 제거.
        if not self._include_full_template:
            try:
                wb.remove(wb[SHEET_FULL_NAME])
            except Exception:
                pass

        self._dst.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self._dst))
        # SharePoint / MIP 메타데이터 제거 — 회사 Excel 에서 ‘읽기 전용’ /
        # ‘보호 보기’ 로 열리는 것을 방지.
        try:
            _strip_corporate_metadata(self._dst)
        except Exception:
            # 메타데이터 정리 실패는 치명적이지 않다 — 결과 파일은 이미 저장됨.
            pass

    # ------------------------------------------------------------------
    def _summary_sheet_name(self) -> str:
        """요약 시트 이름 = 결과 파일명(확장자 제외).  엑셀 시트명 제약(31자·금지문자)
        에 맞춰 정리하고, 비면 안전한 기본값을 쓴다."""
        import re as _re
        name = Path(self._dst).stem or "결과"
        name = _re.sub(r'[:\\/?*\[\]]', "_", name)   # 엑셀 시트명 금지문자 → _
        name = name.strip() or "결과"
        if name == SHEET_FULL_NAME:                  # 2번째 시트와 충돌 방지
            name = name + " "
        return name[:31]

    def _build_summary_sheet(self, wb, rows_input: list, *, index: int = 0) -> None:
        """요약 시트 — A~D 열만, 전체 데이터."""
        self._build_ad_sheet(wb, self._summary_sheet_name(), index, rows_input)

    def _write_unmatched_sheet(self, wb, unmatched_rows: list) -> None:
        """미매칭 사진만 모은 A~D 시트(이미지 포함) — 첫 번째 시트(index 0)."""
        self._build_ad_sheet(wb, i18n.KO.SHEET_UNMATCHED, 0, unmatched_rows)

    def _build_ad_sheet(self, wb, title: str, index: int, rows_input: list) -> None:
        """A~D(번호·slot·기준/검증 이미지) 전용 시트를 만든다.

        전체 양식 시트의 A~D 헤더/병합/폭/행높이를 복사하고 ``rows_input`` 으로
        이미지를 임베드한다.  요약(전체)·미매칭(부분) 시트가 공유한다."""
        full = wb[SHEET_FULL_NAME]
        ws = wb.create_sheet(title=title, index=index)

        # A~D 헤더(row 1~2) 값/서식 복사.  E~H 는 만들지 않는다(요약 시트엔 없음).
        from copy import copy as _copy
        for r in (1, 2):
            for col in ("A", "B", "C", "D"):
                src = full[f"{col}{r}"]
                dst = ws[f"{col}{r}"]
                dst.value = src.value
                if src.has_style:
                    dst.font = _copy(src.font)
                    dst.fill = _copy(src.fill)
                    dst.border = _copy(src.border)
                    dst.alignment = _copy(src.alignment)
                    dst.number_format = src.number_format
        # 병합 헤더(A1:A2, B1:B2, C1:D1) 재현 — A~D 범위만.
        for rng in ("A1:A2", "B1:B2", "C1:D1"):
            try:
                ws.merge_cells(rng)
            except Exception:
                pass
        # 열 폭 — A~D 만 전체 시트와 동일하게.
        for col in ("A", "B", "C", "D"):
            w = full.column_dimensions[col].width
            if w:
                ws.column_dimensions[col].width = w
        # 데이터 행 높이도 동일하게(이미지가 같은 크기로 들어가도록).
        h = full.row_dimensions[DATA_START_ROW].height or ROW_HEIGHT_PT
        ws.row_dimensions[DATA_START_ROW].height = h
        # A~D 데이터 채우기(이미지는 mid 캐시에서 다시 임베드 — 시트 간 공유 불가).
        self._fill_rows(ws, rows_input)
        data_end = DATA_START_ROW + len(rows_input) - 1
        for rr in range(max(data_end + 1, DATA_START_ROW), ws.max_row + 1):
            a = ws.cell(row=rr, column=1)
            if isinstance(a.value, (int, float)):
                a.value = None

    # ------------------------------------------------------------------
    @staticmethod
    def _ensure_width(ws, col_letter: str, min_w: float) -> None:
        cur = ws.column_dimensions[col_letter].width
        if not cur or cur < min_w:
            ws.column_dimensions[col_letter].width = min_w

    @staticmethod
    def _equalize_column_group(ws, cols: list[str], floor: float) -> None:
        """주어진 열들의 width 를 모두 같은 값으로 통일.

        target = max(현재 지정된 width 중 최대, floor). 모든 입력 컬럼이
        target 으로 설정되어 D == C, E == F == G == H 가 보장됨 (#3).
        """
        widths: list[float] = []
        for c in cols:
            w = ws.column_dimensions[c].width
            if w:
                widths.append(float(w))
        target = max(widths + [float(floor)])
        for c in cols:
            # ColumnDimension.customWidth 는 property (no setter) — width 만
            # 세팅하면 openpyxl 이 자동으로 customWidth=True 처리.
            ws.column_dimensions[c].width = target

    @staticmethod
    def _mirror_paired_column_widths(ws) -> None:
        """병합된 헤더 (예: C1:D1) 의 오른쪽 컬럼이 width 미지정인 경우 왼쪽
        컬럼의 width 를 그대로 복사한다.  양식.xlsx 처럼 ‘왼쪽만 폭 지정’ 한
        템플릿에서 오른쪽 셀이 좁아 사진이 작게 임베드되는 문제 해결."""
        from openpyxl.utils import column_index_from_string, get_column_letter
        for rng in list(ws.merged_cells.ranges):
            # 헤더 행에 걸친 가로 병합만 대상 (단일 행, 가로 폭 ≥ 2)
            if rng.min_row != rng.max_row:
                continue
            if rng.max_col - rng.min_col < 1:
                continue
            left = get_column_letter(rng.min_col)
            left_w = ws.column_dimensions[left].width
            if not left_w:
                continue
            for c in range(rng.min_col + 1, rng.max_col + 1):
                col_letter = get_column_letter(c)
                cd = ws.column_dimensions[col_letter]
                if not cd.width:
                    cd.width = left_w

    # ------------------------------------------------------------------
    def _build_minimal_headers(self, ws) -> None:
        """양식.xlsx 가 없을 때 사용할 최소 헤더 (양식의 구조를 흉내)."""
        from openpyxl.styles import Alignment, Font, PatternFill
        yellow = PatternFill("solid", fgColor="FFFF00")
        center = Alignment(horizontal="center", vertical="center")
        ws["A1"] = "No"
        ws["B1"] = "slot#"
        ws["C1"] = "Scan Defect"
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:D1")
        for coord in ("A1", "B1", "C1"):
            c = ws[coord]
            c.font = Font(bold=True)
            c.fill = yellow
            c.alignment = center
        # row 2 의 AOI-N 자리는 _do_export 에서 채움.
        for coord in ("C2", "D2"):
            c = ws[coord]
            c.font = Font(bold=True)
            c.fill = yellow
            c.alignment = center
        ws.row_dimensions[1].height = 21.75
        ws.row_dimensions[2].height = 19.5

    # ------------------------------------------------------------------
    def _write_slot_cell(self, ws, row: int, slot: str, center) -> None:
        """B열에 slot명을 쓴다.  KLA 장비면 slot명(WaferID) 아래 줄에 KLA 하위폴더명을
        **회색 글씨**로 함께 표기한다 (#KLA).  rich text 미지원 시 plain 폴백."""
        from openpyxl.styles import Alignment

        cell = ws[f"{COL_SLOT}{row}"]
        kf = (self._result.kla_folders or {}).get(slot)
        if not kf:
            cell.value = slot
            cell.alignment = center
            return
        wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            cell.value = CellRichText(
                TextBlock(InlineFont(), f"{slot}\n"),
                TextBlock(InlineFont(sz=8, color="808080"), str(kf)),
            )
        except Exception:
            cell.value = f"{slot}\n{kf}"
        cell.alignment = wrap

    # ------------------------------------------------------------------
    def _embed_image_path(self, src: Path) -> Path:
        """셀에 임베드할 이미지 경로를 고른다.

        원본 화질 옵션이 켜져 있으면 원본 파일을 그대로 쓰고(축소 없음),
        꺼져 있으면 중간 화질 캐시(`get_mid_path`)를 쓴다 — 기본은 가볍고 빠름.
        표시 크기는 어느 쪽이든 ``_fit_to_cell`` 이 셀에 맞게 줄이므로, 차이는
        '저장되는 픽셀 데이터의 해상도'(=화질)뿐이다.
        """
        if self._original_quality:
            return Path(src)
        return image_io.get_mid_path(Path(src))

    # ------------------------------------------------------------------
    def _fill_rows(self, ws, rows_input: list[tuple[str, str, object]]) -> None:
        from openpyxl.comments import Comment
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Border, Font, Side

        total = len(rows_input)
        row = DATA_START_ROW
        red_font = Font(color="FFFF2D55", bold=True)
        center = Alignment(horizontal="center", vertical="center")
        # 슬롯이 바뀌는 첫 행 위에 굵은 가로 구분선 (#4).  같은 슬롯끼리
        # 시각적으로 묶이도록.
        slot_sep_side = Side(border_style="thick", color="FF333333")
        prev_slot: Optional[str] = None
        # 템플릿 데이터 행의 ‘기준 높이’ 를 한 번만 측정 — 보통 165.75pt.
        # 양식이 없거나 데이터 행에 높이가 안 잡혀 있으면 ROW_HEIGHT_PT 사용.
        template_row_h = ws.row_dimensions[DATA_START_ROW].height or ROW_HEIGHT_PT
        # C / D 컬럼 폭은 행마다 동일하므로 한 번만 계산.
        cell_w_px = _col_width_to_px(
            ws.column_dimensions[COL_REF].width or IMG_COL_WIDTH
        )
        cell_h_px = _row_height_to_px(template_row_h)
        for idx, (cur_slot, _key, payload) in enumerate(rows_input, start=1):
            # 새 행은 템플릿의 데이터 행과 같은 높이로 통일 → 양식 안팎 일관성.
            cur_h = ws.row_dimensions[row].height
            if not cur_h or cur_h < template_row_h:
                ws.row_dimensions[row].height = template_row_h

            # 슬롯 변경 시 A~H 전 열에 top border 적용 (기존 좌/우/하 보존).
            if prev_slot is not None and cur_slot != prev_slot:
                for col in BORDER_COLS:
                    cell = ws[f"{col}{row}"]
                    old = cell.border
                    cell.border = Border(
                        top=slot_sep_side,
                        left=old.left, right=old.right, bottom=old.bottom,
                        diagonal=old.diagonal,
                        diagonal_direction=old.diagonal_direction,
                        outline=old.outline,
                        vertical=old.vertical,
                        horizontal=old.horizontal,
                    )
            prev_slot = cur_slot

            # A 열: 행 번호 (사용자 양식의 ‘No’).
            no_cell = ws[f"{COL_NO}{row}"]
            no_cell.value = idx
            no_cell.alignment = center

            if isinstance(payload, MatchResult):
                m = payload
                self._write_slot_cell(ws, row, m.slot, center)
                # 손상/누락 이미지 1 장 때문에 전체 export 가 abort 되지 않도록
                # 각 사진을 개별 try 로 감싼다 (Bug #3).  실패하면 파일명 텍스트
                # 로 대체하고 이어서 진행.
                try:
                    ref_mid = self._embed_image_path(m.ref_path)
                    xli_ref = XLImage(str(ref_mid))
                    _fit_to_cell(xli_ref, cell_w_px, cell_h_px)
                    _add_image_centered(ws, xli_ref, COL_REF, row,
                                        cell_w_px, cell_h_px)
                except Exception:
                    ws[f"{COL_REF}{row}"] = str(Path(m.ref_path).name)
                    ws[f"{COL_REF}{row}"].alignment = center
                try:
                    val_mid = self._embed_image_path(m.val_path)
                    xli_val = XLImage(str(val_mid))
                    _fit_to_cell(xli_val, cell_w_px, cell_h_px)
                    _add_image_centered(ws, xli_val, COL_VAL, row,
                                        cell_w_px, cell_h_px)
                except Exception:
                    ws[f"{COL_VAL}{row}"] = str(Path(m.val_path).name)
                    ws[f"{COL_VAL}{row}"].alignment = center
                self.signals.progress.emit(idx, total, m.slot)
            else:
                u: MissEntry = payload
                self._write_slot_cell(ws, row, u.slot, center)
                # 기준 이미지: 정상 임베드.
                try:
                    ref_mid = self._embed_image_path(Path(u.path))
                    xli_ref = XLImage(str(ref_mid))
                    _fit_to_cell(xli_ref, cell_w_px, cell_h_px)
                    _add_image_centered(ws, xli_ref, COL_REF, row,
                                        cell_w_px, cell_h_px)
                except Exception:
                    ws[f"{COL_REF}{row}"] = str(Path(u.path).name)
                # 검증 컬럼에 파일명 텍스트 (빨강).  결함 geometry(area/width/
                # length/contrast) 또는 명시적 마커를 파일명 아래 회색으로 덧붙인다
                # (#geometry).  geometry 비활성(스키마 미충전) 이면 기존과 동일.
                cell_val = ws[f"{COL_VAL}{row}"]
                name = Path(u.path).name
                # geometry(Surface.flt) + 좌표(col/row/x/y, 매칭단계 메커니즘 재사용).
                # 좌표는 Surface.flt 유무와 무관하므로 미지원 자재 행에도 붙는다.
                blocks = self._geometry_blocks(u.path) + self._coord_blocks(u.path)
                if blocks:
                    from openpyxl.cell.rich_text import CellRichText, TextBlock
                    from openpyxl.cell.text import InlineFont
                    red_inline = InlineFont(b=True, color="FFFF2D55")
                    cell_val.value = CellRichText(
                        TextBlock(red_inline, name), *blocks,
                    )
                else:
                    cell_val.value = name
                    cell_val.font = red_font
                cell_val.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True,
                )
                cell_val.comment = Comment("미매칭", "AOI")
                self.signals.progress.emit(idx, total, u.slot)

            row += 1

    # ------------------------------------------------------------------
    @staticmethod
    def _geometry_blocks(path) -> list:
        """미매칭 행 D열 파일명 아래에 덧붙일 회색 TextBlock 목록.

        Surface.flt 에서 결함 geometry 를 환산해 area/width/length/contrast 를
        표기하고, 값을 못 얻으면 '명시적 마커'(미지원 자재 / 데이터 없음)를 넣는다.
        geometry 비활성(스키마 미충전, status="disabled") 이면 빈 목록 → 호출부가
        기존과 동일하게 plain 파일명을 쓴다.  best-effort — 절대 raise 안 함.
        """
        try:
            from ..coords import geometry
            from openpyxl.cell.rich_text import TextBlock
            from openpyxl.cell.text import InlineFont

            res = geometry.resolve(Path(path))
            if res.status == "disabled":
                return []
            grey = InlineFont(sz=8, color="FF808080")
            if res.status == "ok" and res.geometry is not None:
                g = res.geometry
                # contrast 는 일부 자재(예: PI)만 측정 — 대부분 0 이다. 0 이면 '측정
                # 안 함'을 뜻하므로 '—' 로 표기(0.00 으로 오해 방지).
                contrast_txt = ("contrast —" if g.contrast == 0
                                else f"contrast {g.contrast:.2f}")
                # 표기 순서: recipe/zone → area → width → length → contrast
                # (이어서 _coord_blocks 가 col/row → x/y 를 덧붙인다).
                # 이름만 표기(코드 숫자 없이). 이름을 못 찾은 자재만 코드로 폴백(빈칸 방지).
                recipe_disp = g.recipe_name or str(g.recipe)
                zone_disp = g.zone_name or str(g.zone)
                return [
                    TextBlock(grey, f"\nrecipe {recipe_disp} / zone {zone_disp}"),
                    TextBlock(grey, f"\narea {g.area_um2:.2f} ㎛²"),
                    TextBlock(grey, f"\nwidth {g.width_um:.2f} ㎛"),
                    TextBlock(grey, f"\nlength {g.length_um:.2f} ㎛"),
                    TextBlock(grey, f"\n{contrast_txt}"),
                ]
            if res.status == "no_flt":
                return [TextBlock(grey, f"\n{i18n.KO.GEOM_NOT_SUPPORTED}")]
            # "no_data" (Surface.flt 는 있으나 매칭 실패)
            return [TextBlock(grey, f"\n{i18n.KO.GEOM_NO_DATA}")]
        except Exception:
            # rich_text 미지원 openpyxl 등 — 마커 없이 기존 동작으로 폴백.
            return []

    # ------------------------------------------------------------------
    @staticmethod
    def _coord_blocks(path) -> list:
        """미매칭 행 D열에 덧붙일 좌표(col/row/x/y) 회색 TextBlock 목록.

        매칭 단계와 **동일한 메커니즘**(:func:`coords.resolve`)을 그대로 써서
        die col/row 와 die 내부 local x/y(µm)를 얻는다 — col/row/x/y 모두
        DefectCoord 에 이미 함께 들어 있다.  Surface.flt 유무와 무관하므로
        측정정보 미지원 자재 행에도 위치 식별용으로 붙는다.  best-effort.
        """
        try:
            from ..coords import resolve as resolve_coord
            from openpyxl.cell.rich_text import TextBlock
            from openpyxl.cell.text import InlineFont

            c = resolve_coord(Path(path))
            if c is None:
                return []
            grey = InlineFont(sz=8, color="FF808080")
            blocks = [
                TextBlock(grey, f"\ncol {c.col} / row {c.row}"),
                TextBlock(grey, f"\nx {c.x:.0f} / y {c.y:.0f} ㎛"),
            ]
            # KLA 결함은 위 변환값(Camtek 좌표계)에 더해 자체 원본 좌표(XREL/YREL)도 표기.
            if (c.source == "kla" and c.native_x is not None
                    and c.native_y is not None):
                blocks.append(TextBlock(grey, i18n.KO.EXPORT_KLA_NATIVE_FMT.format(
                    x=c.native_x, y=c.native_y)))
            return blocks
        except Exception:
            return []

    # ------------------------------------------------------------------
    def _write_slot_mismatch_sheet(self, wb) -> None:
        ws = wb.create_sheet(title=i18n.KO.SLOT_MISMATCH_SHEET)
        ws["A1"] = "구분"
        ws["B1"] = "Slot 명"
        r = 2
        for s in self._result.slot_only_ref:
            ws.cell(row=r, column=1, value="기준 전용")
            ws.cell(row=r, column=2, value=s)
            r += 1
        for s in self._result.slot_only_val:
            ws.cell(row=r, column=1, value="검증 전용")
            ws.cell(row=r, column=2, value=s)
            r += 1


# ---------------------------------------------------------------------------
# 사진 ↔ 셀 크기 정합 헬퍼
# ---------------------------------------------------------------------------
# Excel 의 column width 는 ‘기본 폰트의 0 자리 글자 수’ 단위라 직접 px 변환이
# 까다롭다.  Calibri 11pt 기준 1 unit ≈ 7 px 정도가 일반 통용 근사값.
# row height 는 pt 단위이므로 96 DPI 환산 (1pt = 4/3 px).
def _col_width_to_px(width_units: float) -> int:
    return max(8, int(round((float(width_units) or 0) * 7.0)))


def _row_height_to_px(height_pt: float) -> int:
    return max(8, int(round((float(height_pt) or 0) * 4.0 / 3.0)))


def _fit_to_cell(xli, cell_w_px: int, cell_h_px: int) -> None:
    """openpyxl 의 Image 를 셀 크기에 ‘비율 유지 + 한쪽 변 가득’ 으로 맞춤.

    가로/세로 중 비율상 먼저 셀에 닿는 변이 cell 의 변 길이에 정확히 일치하고
    반대 변은 남는 여백이 생긴다 (사용자 요청: ‘가로나 세로가 셀에 딱 들어맞을
    때까지 크게’).
    """
    try:
        w = float(xli.width)
        h = float(xli.height)
    except Exception:
        return
    if w <= 0 or h <= 0:
        return
    scale = min(cell_w_px / w, cell_h_px / h)
    if scale <= 0:
        return
    xli.width = max(1, int(round(w * scale)))
    xli.height = max(1, int(round(h * scale)))


def _add_image_centered(ws, xli, col_letter: str, row: int,
                        cell_w_px: int, cell_h_px: int) -> None:
    """``_fit_to_cell`` 로 맞춘 이미지를 셀 안에 **중앙 정렬**로 삽입.

    openpyxl 기본 동작은 셀 좌상단 고정이라 비율상 남는 여백이 한쪽(우/하)에
    몰려 사진이 작아 보인다.  ``OneCellAnchor`` + 오프셋으로 남는 여백을 양쪽에
    균등 분배해 시각적으로 ‘셀에 가득’ 차도록 한다 (크롭/왜곡 없음).
    """
    from openpyxl.drawing.spreadsheet_drawing import (AnchorMarker,
                                                      OneCellAnchor)
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils import column_index_from_string
    from openpyxl.utils.units import pixels_to_EMU

    img_w = int(xli.width)
    img_h = int(xli.height)
    x_off = max(0, (cell_w_px - img_w) // 2)
    y_off = max(0, (cell_h_px - img_h) // 2)
    marker = AnchorMarker(
        col=column_index_from_string(col_letter) - 1,
        colOff=pixels_to_EMU(x_off),
        row=int(row) - 1,
        rowOff=pixels_to_EMU(y_off),
    )
    ext = XDRPositiveSize2D(pixels_to_EMU(img_w), pixels_to_EMU(img_h))
    xli.anchor = OneCellAnchor(_from=marker, ext=ext)
    ws.add_image(xli)


# ---------------------------------------------------------------------------
# SharePoint / MIP 메타데이터 청소
# ---------------------------------------------------------------------------
# 양식.xlsx 가 SharePoint 에서 다운로드된 파일이라 다음 메타데이터를 가지고
# 있다 — 회사 Excel 이 이 파일을 ‘기밀/보호 보기/읽기 전용’ 으로 여는 원인:
#
#   - customXml/*               : SharePoint Media Service 메타데이터
#   - docMetadata/LabelInfo.xml : Microsoft Information Protection 라벨
#   - docProps/custom.xml       : ContentTypeId 등 SharePoint content type 바인딩
#
# 저장 직후 zip 안에서 이 항목들을 제거하고, 참조하는 [Content_Types].xml /
# _rels 파일에서도 해당 라인을 삭제한다.
import re as _re
import zipfile as _zip

_STRIP_PREFIXES = ("customXml/", "docMetadata/")
_STRIP_REL_TARGETS = _re.compile(
    r'(?i)Target="(?:[^"]*/)?(?:customXml|docMetadata)[^"]*"'
)
_STRIP_CONTENT_TYPE_OVERRIDES = _re.compile(
    r'<Override[^>]*PartName="/(?:customXml|docMetadata)[^"]*"[^>]*/>'
)
_STRIP_RELATIONSHIP_LINE = _re.compile(
    r'<Relationship[^/]*?Target="(?:[^"]*/)?(?:customXml|docMetadata)[^"]*"[^/]*?/>'
)


def _strip_corporate_metadata(xlsx_path: Path) -> None:
    """저장된 xlsx 에서 SharePoint / MIP 메타데이터를 제거한다.

    실패해도 결과 파일 자체는 손상되지 않도록 임시 파일에 다시 쓴 뒤 atomic
    rename 으로 교체한다.
    """
    xlsx_path = Path(xlsx_path)
    tmp_out = xlsx_path.with_suffix(xlsx_path.suffix + ".clean.tmp")

    with _zip.ZipFile(xlsx_path, "r") as src:
        names = src.namelist()
        has_metadata = any(
            n.startswith(_STRIP_PREFIXES) for n in names
        )
        if not has_metadata:
            return        # 청소할 게 없으면 그대로 둠

        with _zip.ZipFile(tmp_out, "w", _zip.ZIP_DEFLATED) as dst:
            for info in src.infolist():
                name = info.filename
                if name.startswith(_STRIP_PREFIXES):
                    continue                # 메타데이터 파일은 통째로 제외
                data = src.read(name)
                # 참조 라인 제거 — text XML 만 정리
                if name in (
                    "[Content_Types].xml",
                    "_rels/.rels",
                    "xl/_rels/workbook.xml.rels",
                ):
                    try:
                        text = data.decode("utf-8")
                    except UnicodeDecodeError:
                        dst.writestr(info, data)
                        continue
                    text = _STRIP_CONTENT_TYPE_OVERRIDES.sub("", text)
                    text = _STRIP_RELATIONSHIP_LINE.sub("", text)
                    data = text.encode("utf-8")
                elif name == "docProps/custom.xml":
                    # ContentTypeId 만 가진 custom.xml 은 통째로 비워도 무방.
                    # Excel 이 ContentTypeId 를 보면 SharePoint 문서로 인식.
                    try:
                        text = data.decode("utf-8")
                        if 'name="ContentTypeId"' in text:
                            # 빈 properties 로 대체.
                            data = (
                                b'<?xml version="1.0" encoding="UTF-8" '
                                b'standalone="yes"?>\n'
                                b'<Properties xmlns='
                                b'"http://schemas.openxmlformats.org/'
                                b'officeDocument/2006/custom-properties"/>'
                            )
                    except UnicodeDecodeError:
                        pass
                dst.writestr(info, data)

    tmp_out.replace(xlsx_path)
