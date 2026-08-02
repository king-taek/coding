"""``dev/양식.xlsx`` (엑셀 출력 템플릿) 를 다시 굽는다.

``make_ir.py`` 와 같은 성격의 내부 도구다 — **저장소에 커밋된 바이너리 산출물**을
사람이 손으로 만든 것이 아니라 이 파일에서 재생성한다.  양식을 바꿀 일이 생기면
아래 상수만 고치고 다시 돌린다::

    python scripts/internal/make_template.py

★ 치수·색의 단일 출처는 여기가 아니라 ``exporter`` 다.  앱이 출력할 때 같은 값을
  써야 하므로 상수를 그쪽에서 가져온다 — 두 벌로 갈라지면 템플릿과 실제 출력이
  달라진다.
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from aoi_verification.app.workers.exporter import (BODY_FILLS, COL_WIDTHS,
                                                   GROUP_FILLS, HEADER_NAVY,
                                                   ROW_HEIGHT_PT, TEMPLATE_FONT)

DST = REPO_ROOT / "dev" / "양식.xlsx"
PRESET_ROWS = 20          # A3..A22 에 1..20 을 미리 박아 두던 기존 관습 유지

THIN = Side(style="thin", color="FFB0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "양식"

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 21.75
    ws.row_dimensions[2].height = 19.5

    for rng in ("A1:A2", "B1:B2", "C1:D1", "E1:F1", "G1:H1"):
        ws.merge_cells(rng)
    ws["A1"] = "No"
    ws["B1"] = "slot#"
    ws["C1"] = "Scan Defect"
    ws["E1"] = "Escape Defect [Camtek]"
    ws["G1"] = "Escape Defect [KLA]"
    # row 2 의 AOI-N 은 출력 시 실제 호기 번호로 치환된다(_do_export).
    ws["C2"] = "AOI-17"
    ws["D2"] = "AOI-23"
    ws["E2"] = "Camtek"
    ws["F2"] = "KLA"
    ws["G2"] = "Camtek"
    ws["H2"] = "KLA"

    # 1행 = 남색 띠(흰 글씨) / 2행 = 그룹색(검은 글씨).
    # A·B 는 1~2행 병합이라 남색 하나로 간다.
    for col in "ABCDEFGH":
        top = ws[f"{col}1"]
        top.fill = PatternFill("solid", fgColor=HEADER_NAVY)
        top.font = Font(name=TEMPLATE_FONT, size=11, bold=True,
                        color="FFFFFFFF")
        top.border = BOX
        top.alignment = CENTER

        sub = ws[f"{col}2"]
        sub.border = BOX
        sub.alignment = CENTER
        if col not in "AB":
            sub.fill = PatternFill("solid", fgColor=GROUP_FILLS[col])
            sub.font = Font(name=TEMPLATE_FONT, size=11, bold=True,
                            color="FF1F2937")

    # 데이터 행 — 미리 칠해 두는 것은 '빈 양식을 그냥 인쇄해 손으로 쓰는' 용도다.
    # 실제 출력은 행 수가 이보다 훨씬 많으므로 exporter 가 행마다 다시 칠한다.
    for i in range(PRESET_ROWS):
        r = 3 + i
        ws.row_dimensions[r].height = ROW_HEIGHT_PT
        ws[f"A{r}"] = i + 1
        for col in "ABCDEFGH":
            c = ws[f"{col}{r}"]
            c.border = BOX
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name=TEMPLATE_FONT, size=11)
            c.fill = PatternFill("solid", fgColor=BODY_FILLS[col][i % 2])

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    return DST


if __name__ == "__main__":
    out = build()
    print(f"[OK] {out.relative_to(REPO_ROOT)} 생성")
